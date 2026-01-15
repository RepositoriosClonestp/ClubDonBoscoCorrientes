"""
Exportador a Excel
Exporta transacciones y reportes financieros
"""

import openpyxl
from openpyxl. styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
from pathlib import Path

from config.settings import CLUB_INFO, EXPORTS_PATH


class ExcelExporter:
    """Clase para exportar datos a Excel"""
    
    def __init__(self):
        self.exports_path = EXPORTS_PATH
    
    def exportar_transacciones(self, transacciones: list, fecha_desde: str, fecha_hasta: str) -> str:
        """
        Exporta transacciones a Excel
        
        Args:
            transacciones: Lista de transacciones
            fecha_desde: Fecha inicio del período
            fecha_hasta:  Fecha fin del período
        
        Returns:
            Nombre del archivo generado
        """
        # Crear workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Transacciones"
        
        # Estilos
        header_fill = PatternFill(start_color="1D71B8", end_color="1D71B8", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Título
        ws. merge_cells('A1:H1')
        ws['A1'] = f"{CLUB_INFO['nombre']} - Reporte de Transacciones"
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # Período
        ws.merge_cells('A2:H2')
        fecha_desde_fmt = datetime.strptime(fecha_desde, '%Y-%m-%d').strftime('%d/%m/%Y')
        fecha_hasta_fmt = datetime.strptime(fecha_hasta, '%Y-%m-%d').strftime('%d/%m/%Y')
        ws['A2'] = f"Período:  {fecha_desde_fmt} - {fecha_hasta_fmt}"
        ws['A2'].alignment = Alignment(horizontal='center')
        
        # Encabezados
        headers = ["Fecha", "Tipo", "Categoría", "Descripción", "Monto", "Método", "Comprobante", "Responsable"]
        ws.append([])  # Fila vacía
        ws.append(headers)
        
        header_row = ws[4]
        for cell in header_row:
            cell.fill = header_fill
            cell. font = header_font
            cell.alignment = Alignment(horizontal='center')
            cell.border = border
        
        # Datos
        total_ingresos = 0
        total_egresos = 0
        
        for trans in transacciones:
            fecha = datetime.strptime(trans['fecha'], '%Y-%m-%d').strftime('%d/%m/%Y')
            ws.append([
                fecha,
                trans['tipo']. capitalize(),
                trans['categoria'],
                trans['descripcion'],
                trans['monto'],
                trans. get('metodo_pago', '-') or '-',
                trans.get('comprobante', '-') or '-',
                trans.get('responsable', '-') or '-'
            ])
            
            if trans['tipo'] == 'ingreso':
                total_ingresos += trans['monto']
            else:
                total_egresos += trans['monto']
            
            # Aplicar bordes
            for cell in ws[ws.max_row]:
                cell.border = border
        
        # Totales
        ws.append([])
        ws.append(["", "", "", "TOTAL INGRESOS:", total_ingresos, "", "", ""])
        ws.append(["", "", "", "TOTAL EGRESOS:", total_egresos, "", "", ""])
        ws.append(["", "", "", "BALANCE:", total_ingresos - total_egresos, "", "", ""])
        
        # Formatear totales
        for row in range(ws.max_row - 2, ws.max_row + 1):
            ws[f'D{row}']. font = Font(bold=True)
            ws[f'E{row}'].font = Font(bold=True)
            ws[f'E{row}'].number_format = '"$"#,##0.00'
        
        # Formato de montos
        for row in range(5, ws.max_row - 3):
            ws[f'E{row}'].number_format = '"$"#,##0.00'
        
        # Ajustar anchos de columna
        ws. column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 10
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 40
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F']. width = 15
        ws. column_dimensions['G'].width = 15
        ws.column_dimensions['H'].width = 20
        
        # Guardar archivo
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"transacciones_{timestamp}.xlsx"
        filepath = self.exports_path / filename
        
        wb.save(str(filepath))
        return filename
    
    def exportar_socios(self, socios: list) -> str:
        """
        Exporta lista de socios a Excel
        
        Args:
            socios:  Lista de socios
        
        Returns:
            Nombre del archivo generado
        """
        wb = openpyxl. Workbook()
        ws = wb.active
        ws.title = "Socios"
        
        # Estilos
        header_fill = PatternFill(start_color="1D71B8", end_color="1D71B8", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Título
        ws.merge_cells('A1:I1')
        ws['A1'] = f"{CLUB_INFO['nombre']} - Lista de Socios"
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # Fecha de generación
        ws.merge_cells('A2:I2')
        ws['A2'] = f"Generado el: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        ws['A2']. alignment = Alignment(horizontal='center')
        
        # Encabezados
        headers = ["DNI", "Apellido", "Nombre", "Categoría", "Teléfono", "Email", "Estado Pago", "Último Pago", "Fecha Inscripción"]
        ws.append([])
        ws.append(headers)
        
        header_row = ws[4]
        for cell in header_row: 
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
            cell.border = border
        
        # Datos
        for socio in socios:
            ultimo_pago = socio. get('fecha_ultimo_pago', '')
            if ultimo_pago:
                try:
                    ultimo_pago = datetime.strptime(ultimo_pago, '%Y-%m-%d').strftime('%d/%m/%Y')
                except: 
                    ultimo_pago = '-'
            else:
                ultimo_pago = '-'
            
            fecha_insc = socio.get('fecha_inscripcion', '')
            if fecha_insc:
                try:
                    fecha_insc = datetime.strptime(fecha_insc, '%Y-%m-%d').strftime('%d/%m/%Y')
                except:
                    fecha_insc = '-'
            else: 
                fecha_insc = '-'
            
            ws.append([
                socio['dni'],
                socio['apellido'],
                socio['nombre'],
                socio['categoria'],
                socio. get('telefono', '-') or '-',
                socio. get('email', '-') or '-',
                socio['estado_pago']. upper(),
                ultimo_pago,
                fecha_insc
            ])
            
            # Aplicar bordes
            for cell in ws[ws.max_row]:
                cell.border = border
        
        # Ajustar anchos
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D']. width = 15
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F']. width = 25
        ws.column_dimensions['G'].width = 15
        ws.column_dimensions['H']. width = 15
        ws.column_dimensions['I'].width = 18
        
        # Guardar
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"socios_{timestamp}.xlsx"
        filepath = self.exports_path / filename
        
        wb.save(str(filepath))
        return filename